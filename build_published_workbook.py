"""
Build a published workbook for each company that has a template registered.

This runs AFTER refresh.py in the GitHub Actions workflow. For each company with
a templates/{cid}_template.xlsx file, this script:

  1. Loads the template
  2. Updates the "Live Data" tab in-place using the just-refreshed observations
  3. Saves as data/{cid}/CoverageX_Model_latest.xlsx (or similar)

Recipients open the URL → always-fresh xlsx → no security prompts, no Excel features
needed, no Python on recipient side. The data is fresh because it was rebuilt server-side
when the cron ran.

URL pattern (stable across runs):
  https://raw.githubusercontent.com/BodeBoyer/live-data/main/data/{cid}/{cid}_latest.xlsx
"""
from __future__ import annotations
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = SCRIPT_DIR / "templates"
DATA_DIR = SCRIPT_DIR / "data"


def update_live_data_tab(workbook_path: Path, livedata: dict, sheet_name: str = "Live Data") -> None:
    """Update the workbook's Live Data tab with fresh observations. Matches FRED rows
    by series_id (col E) and Yahoo peer rows by ticker (col A) — same logic as the
    refresh_live_data.py consumer script in the CoverageX repo."""
    wb = openpyxl.load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        return  # template doesn't have a Live Data tab — skip
    ws = wb[sheet_name]

    # Build indexes of EXISTING rows by their natural keys
    fred_rows: dict[str, int] = {}
    yahoo_rows: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        col_a = ws.cell(r, 1).value
        col_d = ws.cell(r, 4).value
        col_e = ws.cell(r, 5).value
        # FRED row: col D == "FRED", col E is series_id (all caps)
        if isinstance(col_e, str) and col_d == "FRED":
            fred_rows[col_e] = r
        # Yahoo peer row: col A is short ticker (1-5 char all uppercase letters), col B is company name
        if isinstance(col_a, str) and 1 <= len(col_a) <= 5 and col_a.isupper() and col_a.isalpha():
            col_b = ws.cell(r, 2).value
            if isinstance(col_b, str) and len(col_b) > 3:
                yahoo_rows[col_a] = r

    # Update header (rows 1-3)
    ws.cell(row=1, column=1).value = f"Live Data — {livedata['display_name']}"
    ws.cell(row=2, column=1).value = (
        f"Auto-refreshed by GitHub Actions at {livedata['refreshed_at']}. "
        f"Re-download from the published URL to get the latest."
    )
    ws.cell(row=3, column=1).value = f"Last refreshed: {livedata['refreshed_at']}"

    # Apply observations
    next_fred_row = max(fred_rows.values(), default=14) + 1 if fred_rows else 15
    for obs in livedata["observations"]:
        if obs.get("source") == "FRED":
            sid = obs.get("series_id")
            if sid in fred_rows:
                r = fred_rows[sid]
                ws.cell(r, 2).value = obs["value"]
                ws.cell(r, 7).value = obs.get("as_of", "")
                ws.cell(r, 8).value = livedata["refreshed_at"]
                if not ws.cell(r, 6).value:
                    ws.cell(r, 6).value = obs.get("url", "")
            else:
                # New FRED series — append below existing block
                r = next_fred_row
                next_fred_row += 1
                ws.cell(r, 1).value = obs["name"]
                ws.cell(r, 2).value = obs["value"]
                ws.cell(r, 3).value = obs.get("unit", "")
                ws.cell(r, 4).value = "FRED"
                ws.cell(r, 5).value = sid
                ws.cell(r, 6).value = obs.get("url", "")
                ws.cell(r, 7).value = obs.get("as_of", "")
                ws.cell(r, 8).value = livedata["refreshed_at"]
                fred_rows[sid] = r
        elif obs.get("source") in ("Yahoo Finance", "Yahoo"):
            ticker = obs["name"].replace("_Price", "")
            if ticker in yahoo_rows:
                r = yahoo_rows[ticker]
                ws.cell(r, 12).value = obs["value"]  # col L = Live Price

    wb.save(workbook_path)


def main() -> int:
    if not TEMPLATES_DIR.exists():
        print(f"No {TEMPLATES_DIR} directory — nothing to publish")
        return 0
    templates = sorted(TEMPLATES_DIR.glob("*_template.xlsx"))
    if not templates:
        print(f"No templates in {TEMPLATES_DIR} — nothing to publish")
        return 0

    print(f"=== publish workbooks: found {len(templates)} template(s) ===")
    for tpl in templates:
        cid = tpl.stem.replace("_template", "")
        livedata_path = DATA_DIR / cid / "livedata.json"
        if not livedata_path.exists():
            print(f"  skip {cid}: no livedata.json (refresh.py hasn't run for this company)")
            continue

        # Load the freshly-refreshed observations
        livedata = json.loads(livedata_path.read_text())

        # Output path: data/{cid}/{cid}_latest.xlsx
        out_path = DATA_DIR / cid / f"{cid}_latest.xlsx"
        shutil.copy(tpl, out_path)

        try:
            update_live_data_tab(out_path, livedata)
            print(f"  ✓ published {out_path} (refreshed at {livedata['refreshed_at']})")
        except Exception as e:
            print(f"  ✗ failed {cid}: {e}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
